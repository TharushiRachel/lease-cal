from openpyxl import load_workbook
import sys

try:
    # Load the workbook
    wb = load_workbook('Equated Lease cal.xlsx', data_only=False)
    
    print(f"Excel file has {len(wb.sheetnames)} sheet(s):")
    for sheet_name in wb.sheetnames:
        print(f"  - {sheet_name}")
    
    print("\n" + "="*80 + "\n")
    
    # Process each sheet
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"Sheet: {sheet_name}")
        print(f"Dimensions: {ws.max_row} rows x {ws.max_column} columns")
        print("\n")
        
        # Get all cells with data
        cells_with_data = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if cell.value is not None or (hasattr(cell, 'data_type') and cell.data_type == 'f'):
                    cells_with_data.append(cell)
        
        if not cells_with_data:
            print("No data found in this sheet.")
        else:
            print("Cell Values and Formulas:")
            print("-" * 80)
            
            # Group by row for better display
            current_row = None
            for cell in sorted(cells_with_data, key=lambda c: (c.row, c.column)):
                if current_row != cell.row:
                    if current_row is not None:
                        print()  # Empty line between rows
                    current_row = cell.row
                
                cell_address = cell.coordinate
                cell_value = cell.value
                has_formula = cell.data_type == 'f' if hasattr(cell, 'data_type') else False
                
                if has_formula:
                    formula = cell.value  # This is the formula string
                    # Get the calculated value (if available)
                    try:
                        # Try to get the calculated value by loading with data_only=True
                        wb_data = load_workbook('Equated Lease cal.xlsx', data_only=True)
                        ws_data = wb_data[sheet_name]
                        calculated_value = ws_data[cell_address].value
                        print(f"{cell_address}: Formula = {formula} | Calculated Value = {calculated_value}")
                    except:
                        print(f"{cell_address}: Formula = {formula} | Value = {cell_value}")
                else:
                    print(f"{cell_address}: Value = {cell_value}")
        
        print("\n" + "="*80 + "\n")
        
except Exception as e:
    print(f"Error reading Excel file: {e}")
    import traceback
    traceback.print_exc()

